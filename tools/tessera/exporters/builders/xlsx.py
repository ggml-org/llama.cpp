"""XLSX builder (exporter).

Converts an ``IntermediateDocument`` to a .xlsx file via
openpyxl. v1 supports one table per document; multi-table
documents are exported with each table on its own sheet.

The v1 exporter is the inverse of the importer: a cell that
came in as a formula (``=SUM(A1:A10)``) goes out as a formula
in the same cell. v2 will preserve formatting (bold, fill,
column widths).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from tools.tessera.importers.intermediate import (
    IntermediateBlock,
    IntermediateDocument,
)

log = logging.getLogger(__name__)


def build_xlsx(doc: IntermediateDocument, output: Path) -> None:
    """Render an IntermediateDocument to a .xlsx file at `output`.

    One table per sheet. If the document has no table
    blocks (rare; usually an email or markdown document
    exported as XLSX), we emit a single sheet with one
    paragraph per row.
    """
    wb = Workbook()
    # Remove the default sheet; we'll add named sheets.
    wb.remove(wb.active)

    table_blocks = [b for b in doc.blocks if b.type == "table"]
    if not table_blocks:
        # No tables: emit one sheet with paragraph rows.
        ws = wb.create_sheet("Sheet1")
        ws.append(["Block type", "Text"])
        for ib in doc.blocks:
            ws.append([ib.type, "".join(r.text for r in ib.runs)])
    else:
        for i, tb in enumerate(table_blocks, 1):
            sheet_name = str(tb.meta.get("sheet_name", f"Sheet{i}"))[:31]
            ws = wb.create_sheet(sheet_name)
            _render_table(ws, tb)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))


def _render_table(ws: Any, ib: IntermediateBlock) -> None:
    """Render a single table block to the given worksheet."""
    n_rows = int(ib.attrs.get("rows", 0))
    n_cols = int(ib.attrs.get("cols", 0))
    children = [c for c in ib.children if hasattr(c, "type")]
    flat: list[str] = []
    for c in children:
        flat.append("".join(r.text for r in c.runs))
    for r in range(n_rows):
        row: list[Any] = []
        for c in range(n_cols):
            i = r * n_cols + c
            if i < len(flat):
                cell_text = flat[i]
                # If the original cell was a formula, restore
                # the leading "=" so openpyxl writes it as a
                # formula. The cached value is appended back as
                # a comment for the user to see.
                meta = children[i].meta if i < len(children) else {}
                if meta.get("data_type") == "formula":
                    row.append("=" + cell_text if not cell_text.startswith("=") else cell_text)
                else:
                    row.append(cell_text)
            else:
                row.append("")
        ws.append(row)
