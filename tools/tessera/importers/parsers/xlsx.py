"""XLSX parser.

Reads a .xlsx with openpyxl and produces a list of
``IntermediateDocument``s (one per sheet). Each sheet is a
``table`` block with cell blocks; the cell content is
the formula (if present) followed by the cached value.

The v1 XLSX importer makes two practical compromises:

* **Formulas are text, not live.** A cell that says
  ``=SUM(A1:A10)`` produces a cell block whose text is
  ``"=SUM(A1:A10) = 55"`` (formula + value). The agent can read
  the formula but the AST doesn't re-evaluate it. Re-evaluating
  formulas is punted to v2 (see the spec's §10.4 "Punted on").

* **Formatting is dropped.** Bold / fill / border / column
  width are not preserved. The agent can re-add formatting
  via chat if needed.

Coverage:

* Single sheet -> one document with one table block.
* Multi-sheet -> one document per sheet, the importer returns a
  list. The CLI writes each sheet as a separate ``graph_entity``
  linked to a parent "workbook" entity (the link is emitted by
  the receipt emitter).
* Empty sheets -> a single empty table block (the user can see
  the sheet exists).
* Strings, numbers, dates, booleans, errors, formulas: all
  emitted as text.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..intermediate import (
    IntermediateBlock,
    IntermediateDocument,
    IntermediateInlineRun,
)

log = logging.getLogger(__name__)


def parse_xlsx(path: Path) -> list[IntermediateDocument]:
    """Parse a .xlsx file into one IntermediateDocument per sheet.

    Reads the workbook with openpyxl's default mode (data only +
    formulas). The cell value is a Python object (str, int,
    float, datetime, bool, None) depending on the cell's data
    type. We stringify it; the AST builder re-parses only the
    ``=``-prefixed strings into ``code`` runs.
    """
    log.debug("parse_xlsx: %s", path)
    wb = load_workbook(filename=str(path), data_only=False)
    docs: list[IntermediateDocument] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        docs.append(_sheet_to_doc(ws, sheet_name))
    return docs


def _sheet_to_doc(ws: Any, sheet_name: str) -> IntermediateDocument:
    """Convert one openpyxl Worksheet to an IntermediateDocument.

    Empty cells become empty cell blocks (the user / agent needs
    to see the cell exists so they don't lose track of indexing).
    The table block carries rows / cols / cells attributes that
    the AST builder copies into the AST.
    """
    # Find the used range. ws.max_row / ws.max_column are 1-based
    # but can be None for an empty sheet.
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        # Empty sheet: emit a single empty table block.
        table = IntermediateBlock(
            type="table",
            attrs={"rows": 0, "cols": 0, "cells": []},
        )
        return IntermediateDocument(
            blocks=[table],
            meta={"sheet_name": sheet_name},
        )

    cell_grid: list[list[IntermediateBlock]] = []
    for r in range(1, max_row + 1):
        row_blocks: list[IntermediateBlock] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_blocks.append(_cell_to_block(cell))
        cell_grid.append(row_blocks)

    flat: list[IntermediateBlock] = [c for row in cell_grid for c in row]
    table = IntermediateBlock(
        type="table",
        attrs={
            "rows": max_row,
            "cols": max_col,
            "cells": [[c.id for c in row] for row in cell_grid],
        },
        children=flat,
    )
    return IntermediateDocument(blocks=[table], meta={"sheet_name": sheet_name})


def _cell_to_block(cell: Any) -> IntermediateBlock:
    """Build an IntermediateBlock for one openpyxl Cell.

    We format the cell's content as a single run. The
    ``data_type`` is one of:

    * ``"s"`` (string)
    * ``"n"`` (number)
    * ``"b"`` (boolean)
    * ``"d"`` (date)
    * ``"f"`` (formula) - the value is the formula's result
      (openpyxl returns the formula when ``data_only=False``)
    * ``"e"`` (error) - the value is an error string
    * ``"empty"`` - cell has no value
    """
    dtype = getattr(cell, "data_type", None)
    value = cell.value

    if value is None or dtype == "empty":
        return IntermediateBlock(type="paragraph", runs=[])

    # openpyxl exposes the formula in cell.value when
    # data_only=False; the cached value is loaded separately via
    # data_only=True. We use the formula version by default (the
    # user wants to see the formula, not the cached result) and
    # append the cached value if available.
    if dtype == "f":
        formula = str(value)
        cached = _cached_value(cell)
        text = f"{formula} = {cached}" if cached is not None else formula
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=text)],
            meta={"data_type": "formula", "formula": formula, "cached": cached},
        )

    if dtype == "b":
        text = "TRUE" if value else "FALSE"
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=text)],
            meta={"data_type": "bool"},
        )

    if dtype == "d":
        # datetime / date
        text = value.isoformat() if hasattr(value, "isoformat") else str(value)
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=text)],
            meta={"data_type": "date"},
        )

    if dtype == "e":
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=str(value))],
            meta={"data_type": "error"},
        )

    if dtype == "n":
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=_format_number(value))],
            meta={"data_type": "number"},
        )

    # Default: string. Cell.value for a string cell is the raw
    # string; if it starts with "=" treat it as a formula (some
    # spreadsheets store formulas as text).
    text = str(value)
    if text.startswith("="):
        return IntermediateBlock(
            type="paragraph",
            runs=[IntermediateInlineRun(text=text)],
            meta={"data_type": "formula", "formula": text},
        )
    return IntermediateBlock(
        type="paragraph",
        runs=[IntermediateInlineRun(text=text)],
        meta={"data_type": "string"},
    )


def _cached_value(cell: Any) -> str | None:
    """Return the cached value for a formula cell, if available.

    openpyxl exposes this via a parallel data-only load. We
    avoid the second workbook open by re-opening just the
    workbook in data-only mode; that's the documented pattern.
    For the v1 importer this is fine; v2 should batch it.
    """
    try:
        from openpyxl import load_workbook as _lw

        wb = _lw(filename=cell.parent.parent.path, data_only=True)  # type: ignore[attr-defined]
        ws = wb[cell.parent.title]
        v = ws.cell(row=cell.row, column=cell.column).value
        if v is None:
            return None
        return _format_number(v) if isinstance(v, (int, float)) else str(v)
    except Exception as e:  # noqa: BLE001
        log.debug("cached_value failed: %s", e)
        return None


def _format_number(v: Any) -> str:
    """Format a number for display.

    Integers print without a decimal point; floats use str()'s
    default which is short and round-trippable for the common
    case. v2 will use locale-aware formatting.
    """
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        # Avoid the "1.0" -> "1" lossy round-trip; use repr() for
        # the unambiguous form, then strip the trailing ".0" for
        # integer-valued floats.
        s = repr(v)
        if s.endswith(".0"):
            return s[:-2]
        return s
    return str(v)
